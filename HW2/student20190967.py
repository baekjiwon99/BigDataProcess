#!/user9/bin/python3

import openpyxl

wb = openpyxl.load_workbook( "student.xlsx" )
ws = wb['Sheet1']


row_id = 1
sumlist = []

for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value * 0.01
		ws.cell(row = row_id, column = 7).value = sum_v
		sumlist.append(sum_v)
	row_id += 1
sumlist.sort(reverse = True)

totalcount = len(sumlist)
acount = int(totalcount * 0.3)
ascore = sumlist[acount - 1]

apcount = int(acount * 0.5)
apscore = sumlist[apcount - 1]

bcount = int(totalcount * 0.7)
bscore = sumlist[bcount - 1]

bpcount = int(acount + (bcount - acount) * 0.5)
bpscore = sumlist[bpcount - 1]

cpcount = int(bcount + (totalcount * 0.3 * 0.5))
cpscore = sumlist[cpcount - 1]

row_id = 1 
for row in ws:	
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 7).value
		grade = "C0"
		if(sum_v >= apscore):
			grade = 'A+'
		elif(sum_v >= ascore):
			grade = 'A0'
		elif(sum_v >= bpscore):
			grade = 'B+'
		elif(sum_v >= bscore):
			grade = 'B0'
		elif(sum_v >= cpscore):
			grade = 'C+'
		ws.cell(row = row_id, column = 8).value = grade
	row_id += 1

wb.save("student.xlsx")
